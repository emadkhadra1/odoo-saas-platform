import base64
import io
import json
import logging
import re

from odoo import models, fields, api
from odoo.exceptions import ValidationError, UserError


def _strip_diacritics(text):
    """إزالة التشكيل العربي (حركات) لتوحيد المقارنة النصية.
    مثال: 'مسلّمة' → 'مسلمة'، 'مُلغاة' → 'ملغاة'
    """
    # نطاق الحركات العربية: U+0610–U+061A و U+064B–U+065F و U+0670
    return re.sub(r'[\u0610-\u061a\u064b-\u065f\u0670]', '', text)

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.warning("openpyxl not installed. Excel import will not work. Install with: pip install openpyxl")
    openpyxl = None

# بادئات أعمدة كيتا المركّبة — تُحذف قبل المطابقة
KEETA_COL_PREFIXES = [
    'فترة الوردية_',
    ' أحجام المهام_',
    'أحجام المهام_',
    'تجربة التوصيل_',
]

DEFAULT_COLUMN_MAPPING = {
    "row_num": ["م", "#", "Row", "row"],
    "rider_account_id": [
        # عام
        "معرف الحساب", "account_id", "rider_account_id", "Account ID", "Rider ID", "rider_id",
        # هنقر ستيشن
        "DA ID", "Driver ID", "Rider Id", "رقم الحساب", "رقم حساب السائق", "رقم حساب",
        "DA Account No", "Account No", "حساب السائق",
        # تويو
        "Rep ID", "Rep Id", "Representative ID",
        # كيتا
        "معرّف السائق", "معرف السائق", "Driver ID (Keeta)", "Rider Account",
    ],
    "rider_name": [
        "الاسم", "اسم الحساب", "rider_name", "Rider Name", "Driver Name", "Name", "name",
        "اسم المندوب", "اسم السائق", "الاسم الكامل",
        # هنقر ستيشن
        "Rider", "DA Name",
        # كيتا — العمود الثالث (الاسم الأخير)
        "اسم السائق (الأخير)",
    ],
    "rider_user": ["مستخدم الحساب", "Account User", "User", "الاسم الحقيقي"],
    "rider_phone": [
        "phone", "rider_phone", "هاتف", "رقم الجوال", "جوال", "Phone", "Mobile",
        "Rider Phone", "رقم الهاتف",
    ],
    "vehicle_type_company": [
        "نوع المركبة من الشركة", "Vehicle Type Company", "vehicle_type", "نوع السيارة",
        "مركبة الشركة", "Vehicle Company",
        # هنقر ستيشن
        "Vehicle Name",
        # تويو
        "Vehicle Type",
        # كيتا
        "نوع المركبة",
    ],
    "vehicle_type_contract": [
        "نوع المركبة حسب العقد", "Vehicle Type Contract",
        "مركبة العقد", "Vehicle Contract",
    ],
    "plate_number": ["اللوحة", "رقم اللوحة", "Plate Number", "plate", "Plate", "لوحة", "Plate No", "لوحة السيارة"],
    "platform_target": [
        "تاجت كيتا", "تارجت كيتا", "Platform Target", "Keeta Target", "KT Target", "تارجيت",
        "تارجت المنصة", "تارجيت المنصة", "Target", "Daily Target", "الهدف اليومي", "الهدف",
    ],
    "company_target_val": [
        "التارجت الشركة", "تارجت الشركة", "Company Target", "CT Target", "هدف الشركة",
    ],
    "accepted_tasks": [
        "المهام المقبولة", "Accepted Tasks", "Accepted", "accepted", "مقبولة",
        "المهام المقبوله", "Accepted Orders",
        # هنقر ستيشن
        "Accepted Deliveries",
        # تويو
        "# Orders Accepted",
    ],
    "delivered_tasks": [
        "المهام التي تم تسليمها", "Delivered Tasks", "Delivered", "delivered",
        "مسلمة", "مسلّمة", "التسليم", "Completed",
        "المهام المسلمة", "المهام المسلّمة", "Deliveries",
        # هنقر ستيشن
        "Completed Deliveries",
        # تويو
        "# Completed Orders",
        # تويو VDA
        "Sum of total delivered tasks", "Total delivered tasks",
        # كيتا
        "المهام التي تم تسليمها",
    ],
    "large_orders_completed": [
        "مهام الطلبات الكبيرة المكتملة", "Large Orders Completed", "Large Orders",
        # كيتا
        "مهام الطلبات الكبيرة المكتملة",
    ],
    "cancelled_tasks": [
        "المهام المُلغاة", "المهام الملغاة", "Cancelled Tasks", "Cancelled", "cancelled",
        "ملغاة", "ملغيات", "المهام الملغيه", "Cancellations",
        # هنقر ستيشن
        "Cancelled Deliveries",
        # تويو (أول عمود إلغاء)
        "# Cancelled Courier Orders",
        # تويو VDA
        "Cancellation(D)", "Cancellation (D)",
    ],
    "rejected_tasks": [
        "المهام المرفوضة", "Rejected Tasks", "Rejected", "rejected", "مرفوضة",
        "مرفوضه", "Rejections",
        # هنقر ستيشن
        "Declined Deliveries",
        # كيتا (بدون بادئة — يُعالج بعد حذف البادئة)
        " المهام المرفوضة",
    ],
    "driver_rejected": [
        "المهام المرفوضة (السائق)", "Driver Rejected", "driver_rejected",
        # كيتا (بدون بادئة)
        "المهام المرفوضة (السائق)",
    ],
    "auto_rejected": [
        "المهام المرفوضة تلقائيًا (تلقائياً)", "المهام المرفوضة تلقائيًا", "Auto Rejected", "auto_rejected",
        # كيتا
        "المهام المرفوضة تلقائيًا (تلقائياً)",
    ],
    "online_hours": [
        "وقت اتصال السائقين عبر تطبيق السائق.", "وقت اتصال السائقين",
        "Online Hours", "online_hours", "Connection Time", "ساعات العمل",
        "ساعات أونلاين", "ساعات الاتصال", "ساعات الاتصال أونلاين",
        "Online Time", "Hours Online", "Logged Hours",
        # هنقر ستيشن
        "Actual Working Hours",
        # تويو
        "Hours Online (Total)",
        # كيتا (بدون بادئة)
        "وقت اتصال السائقين عبر تطبيق السائق.",
        "courier_valid_online_duration",
    ],
    "peak_hours": [
        "ساعات الذروة", "Peak Hours", "peak_hours", "Piek Hours", "أوقات الذروة",
        # كيتا (بدون بادئة)
        "ساعات الاتصال في وقت الذروة",
    ],
    "ontime_rate": [
        "نسبة الطلبات التي تم تسليمها في الوقت المحدد (D)", "نسبة التسليم في الوقت",
        "On-time Rate", "ontime_rate", "OTD",
        # كيتا (بدون بادئة)
        "نسبة الطلبات التي تم تسليمها في الوقت المحدد (D)",
    ],
    "large_ontime_rate": [
        "معدل توصيل الطلبات الكبيرة في الوقت المُحدَّد", "Large Order On-time", "large_ontime_rate",
    ],
    "avg_delivery_duration": [
        "متوسط مدة التوصيل لكل طلب مكتمل", "متوسط مدة التوصيل", "Avg Duration", "avg_duration",
    ],
    "over_55min_rate": [
        "نسبة الطلبات المُسلمة (أكثر من 55 دقيقة).", "نسبة أكثر من 55 دقيقة", "Over 55min Rate",
    ],
    "late_tasks": [
        "مهام الطلبات المتأخرة", "Late Tasks", "late_tasks",
        # تويو/كيتا VDA
        "Delivered orders over 55min", "Delivered orders over 55 min",
        "Orders over 55min", "Over 55min orders",
    ],
    "very_late_tasks": [
        "مهام الطلبات المتأخرة جدًا", "Very Late Tasks", "very_late_tasks",
    ],
    "on_time_deliveries": [
        "عدد التسليمات في الوقت", "On-Time Count", "on_time_deliveries",
        # ToYou VDA
        "Sum of on time D tasks", "On Time D", "On-time D tasks",
    ],
    "advance_deliveries": ["عدد التسليمات المسبقة", "Advance Count", "advance_deliveries"],
    "fuel": ["بنزين", "Fuel", "fuel", "Gas", "قيمة البنزين", "مبلغ البنزين"],
    "order_id": ["order_id", "order id", "رقم الطلب", "Order ID", "Order No"],
    "city_name": ["city", "city_name", "المدينة", "City", "City Name"],
    "order_date": [
        "date", "order_date", "التاريخ", "تاريخ الطلب", "Order Date",
        # تويو
        "Day",
        # تويو VDA
        "Online Day",
    ],
    "distance": [
        "distance", "distance_km", "المسافة", "Distance", "Distance (km)", "KM",
        # تويو VDA
        "Total Delivery Distance (km)", "Payable distance",
    ],
    "platform_amount": ["amount", "platform_amount", "المبلغ", "مبلغ المنصة", "Amount", "Platform Amount", "Total"],
    # Monthly summary financial fields
    "capacity_incentive": ["حافز السعة", "Capacity Incentive", "capacity_incentive", "Valid DA Incentive"],
    "experience_incentive": ["حافز الخبرة", "Experience Incentive", "experience_incentive"],
    "subsidy": ["إعانة", "Subsidy", "subsidy", "دعم"],
    "dxg": ["DXG", "dxg"],
    "tips_excl_vat": ["بقشيش", "Tips", "tips_excl_vat", "Tips excl. VAT", "بقشيش بدون ضريبة"],
    "other_activities": ["أنشطة أخرى", "Other Activities", "other_activities", "مكافآت"],
    "deductions": ["خصومات", "Deductions", "deductions"],
    "food_damage_compensation": ["تعويض تلف طعام", "Food Damage", "food_damage_compensation"],
    "other_adjustment": ["تعديل آخر", "Other Adjustment", "other_adjustment"],
    "valid_days": ["الأيام الصالحة", "Valid Days", "valid_days", "Working Days"],
}


class DeliveryImportSession(models.Model):
    _name = 'delivery.import.session'
    _description = 'Delivery Import Session'
    _inherit = ['mail.thread']
    _order = 'import_date desc'

    company_id = fields.Many2one('delivery.company', string='Company', required=True, tracking=True, ondelete='cascade')
    branch_id = fields.Many2one('delivery.company.branch', string='Branch', tracking=True,
                                 domain="[('company_id', '=', company_id)]")
    contract_id = fields.Many2one('delivery.contract', string='Contract', tracking=True,
                                  domain="[('company_id', '=', company_id), ('status', '=', 'active'), '|', ('branch_id', '=', branch_id), ('branch_id', '=', False)]")
    file_name = fields.Char(string='File Name', tracking=True)
    file_data = fields.Binary(string='Excel File', attachment=True)
    import_date = fields.Date(string='Import Date', required=True, default=fields.Date.today, tracking=True)
    period_start = fields.Date(string='Period Start', required=True)
    period_end = fields.Date(string='Period End', required=True)
    sheet_name = fields.Char(string='Sheet Name (اسم الورقة)',
                             help='اسم الورقة المراد قراءتها. اتركه فارغاً لقراءة الورقة الأولى.\n'
                                  'Sheet name to read. Leave empty to read the first sheet.')
    import_all_sheets = fields.Boolean(
        string='📋 استيراد كل الأوراق (كل الأيام)',
        default=False,
        help='عند التفعيل، يُقرأ جميع الأوراق في الملف ويُجمع البيانات.\n'
             'مناسب لملفات تويو وهنقر التي تحتوي ورقة لكل يوم (1، 2، 3 ... 30).\n'
             'اسم الورقة (1-31) + تاريخ بداية الفترة = تاريخ اليوم تلقائياً.\n'
             'When enabled, reads ALL sheets and aggregates rows.\n'
             'Suitable for ToYou / HungerStation files with one sheet per day.'
    )
    auto_create_riders = fields.Boolean(
        string='🚴 إنشاء رايدرز تلقائياً',
        default=False,
        help='عند التفعيل: إذا لم يُعثر على الرايدر بمعرّفه في النظام، يُنشأ تلقائياً باستخدام بيانات الملف.\n'
             'مناسب لملفات كيتا / VDA عند أول رفع للشركة.\n'
             'When enabled: riders not found in the system are created automatically from the file data.\n'
             'Useful for first-time Keeta / VDA file imports.'
    )
    status = fields.Selection([
        ('pending', 'Pending'),
        ('validated', 'Validated'),
        ('failed', 'Failed'),
        ('imported', 'Imported'),
    ], string='Status', default='pending', required=True, tracking=True)
    total_rows = fields.Integer(string='Total Rows', default=0)
    valid_rows = fields.Integer(string='Valid Rows', default=0)
    error_rows = fields.Integer(string='Error Rows', default=0)
    total_amount = fields.Float(string='Total Amount', digits=(12, 2), default=0.0)
    import_type = fields.Selection([
        ('daily', 'أداء يومي (Daily Performance)'),
        ('monthly', 'أداء شهري - ملخص (Monthly Summary)'),
        ('orders', 'تفاصيل طلبات (Order Details)'),
        ('fuel', 'تقرير بنزين يومي (Daily Fuel Report)'),
    ], string='Import Type (نوع الاستيراد)', default='daily', required=True, tracking=True,
        help='daily: صف لكل مندوب في اليوم\nmonthly: صف لكل مندوب في الشهر (ملخص كلي)\n'
             'orders: صف لكل طلب\nfuel: تقرير بنزين يومي (م، الاسم، اللوحة، المبلغ)')
    imported_by = fields.Many2one('res.users', string='Imported By', default=lambda self: self.env.user)
    notes = fields.Text(string='Notes')
    error_log = fields.Text(string='Error Log')

    row_ids = fields.One2many('delivery.import.row', 'session_id', string='Import Rows')
    row_count = fields.Integer(compute='_compute_row_count')

    def _compute_row_count(self):
        for rec in self:
            rec.row_count = len(rec.row_ids)

    @api.onchange('company_id')
    def _onchange_company_id(self):
        if self.company_id:
            self.branch_id = False
            active_contract = self.env['delivery.contract'].search([
                ('company_id', '=', self.company_id.id),
                ('status', '=', 'active'),
            ], limit=1)
            self.contract_id = active_contract.id if active_contract else False

    @api.onchange('branch_id')
    def _onchange_branch_id(self):
        if self.branch_id:
            active_contract = self.env['delivery.contract'].search([
                ('company_id', '=', self.company_id.id),
                ('branch_id', '=', self.branch_id.id),
                ('status', '=', 'active'),
            ], limit=1)
            if active_contract:
                self.contract_id = active_contract.id

    def _get_column_mapping(self):
        """Build a {field_name: [alias, ...]} dict from defaults + contract-specific maps."""
        self.ensure_one()
        mapping = {k: list(v) for k, v in DEFAULT_COLUMN_MAPPING.items()}

        if self.contract_id:
            col_maps = self.env['delivery.column.map'].search([
                ('contract_id', '=', self.contract_id.id),
                ('system_field', '!=', False),
                ('system_field', '!=', '_ignore_'),
                ('system_field', '!=', '_extra_'),
            ])
            for cm in col_maps:
                field = cm.system_field
                if field not in mapping:
                    mapping[field] = []
                if cm.excel_column not in mapping[field]:
                    mapping[field].insert(0, cm.excel_column)

        return mapping

    def _get_contract_extra_map(self):
        """Return {excel_column_lower: custom_key} for columns mapped as _extra_."""
        self.ensure_one()
        result = {}
        if not self.contract_id:
            return result
        col_maps = self.env['delivery.column.map'].search([
            ('contract_id', '=', self.contract_id.id),
            ('system_field', '=', '_extra_'),
        ])
        for cm in col_maps:
            key = cm.custom_key or cm.excel_column.strip().replace(' ', '_').lower()
            result[cm.excel_column.strip().lower()] = key
        return result

    def _get_contract_ignored_cols(self):
        """Return a set of excel_column_lower strings that should be ignored."""
        self.ensure_one()
        result = set()
        if not self.contract_id:
            return result
        col_maps = self.env['delivery.column.map'].search([
            ('contract_id', '=', self.contract_id.id),
            ('system_field', '=', '_ignore_'),
        ])
        for cm in col_maps:
            result.add(cm.excel_column.strip().lower())
        return result

    def _strip_keeta_prefix(self, header_lower):
        """حذف بادئات أعمدة كيتا المركّبة للحصول على الاسم الفعلي للعمود.
        مثال: 'فترة الوردية_وقت اتصال السائقين...' → 'وقت اتصال السائقين...'
        """
        for prefix in KEETA_COL_PREFIXES:
            if header_lower.startswith(prefix.lower()):
                return header_lower[len(prefix):]
        return header_lower

    def _match_header(self, header_str, header_lower, mapping):
        """محاولة مطابقة عنوان عمود مع الـ mapping المعطى.
        يجرّب المطابقة المباشرة أولاً، ثم بعد حذف التشكيل، ثم بعد حذف بادئات كيتا.
        يُعيد اسم الحقل أو None.
        """
        header_clean = _strip_diacritics(header_lower)
        # نسخة منزوعة البادئة (لكيتا)
        stripped_lower = self._strip_keeta_prefix(header_lower)
        stripped_clean = _strip_diacritics(stripped_lower)

        for field_name, aliases in mapping.items():
            alias_list = aliases if isinstance(aliases, list) else [aliases]
            for alias in alias_list:
                alias_str = str(alias).strip()
                alias_lower = alias_str.lower()
                alias_clean = _strip_diacritics(alias_lower)
                # 1. مطابقة مباشرة
                if header_lower == alias_lower or header_str == alias_str:
                    return field_name
                # 2. بعد إزالة التشكيل
                if header_clean == alias_clean:
                    return field_name
                # 3. بعد حذف بادئة كيتا
                if stripped_lower == alias_lower or stripped_clean == alias_clean:
                    return field_name
        return None

    def _detect_columns(self, headers_combined):
        """Return (detected, extra_cols) where:
        - detected = {field_name: col_idx} for known system fields
        - extra_cols = {col_idx: custom_key} for extra/unmapped columns to preserve
        """
        mapping = self._get_column_mapping()
        extra_map = self._get_contract_extra_map()
        ignored = self._get_contract_ignored_cols()

        detected = {}
        extra_cols = {}
        unmapped_headers = []

        for col_idx, header_val in enumerate(headers_combined):
            if not header_val:
                continue
            header_str = str(header_val).strip()
            header_lower = header_str.lower()

            if header_lower in ignored:
                continue

            matched_field = None
            for field_name, aliases in mapping.items():
                if field_name in detected:
                    continue
                # استخدام _match_header مع هذا الحقل وقائمة aliases فقط
                alias_list = aliases if isinstance(aliases, list) else [aliases]
                header_clean = _strip_diacritics(header_lower)
                stripped_lower = self._strip_keeta_prefix(header_lower)
                stripped_clean = _strip_diacritics(stripped_lower)

                for alias in alias_list:
                    alias_str = str(alias).strip()
                    alias_lower = alias_str.lower()
                    alias_clean = _strip_diacritics(alias_lower)
                    if (header_lower == alias_lower or header_str == alias_str
                            or header_clean == alias_clean
                            or stripped_lower == alias_lower
                            or stripped_clean == alias_clean):
                        matched_field = field_name
                        break
                if matched_field:
                    break

            if matched_field:
                detected[matched_field] = col_idx
            else:
                if header_lower in extra_map:
                    extra_cols[col_idx] = extra_map[header_lower]
                else:
                    safe_key = header_str.replace(' ', '_').replace('(', '').replace(')', '').replace('.', '')
                    extra_cols[col_idx] = safe_key
                    unmapped_headers.append(header_str)

        return detected, extra_cols, unmapped_headers

    def action_detect_columns(self):
        """Read file headers and auto-populate delivery.column.map for this contract."""
        self.ensure_one()
        if not self.file_data:
            raise UserError('الرجاء رفع ملف Excel أولاً / Please upload a file first.')
        if not self.contract_id:
            raise UserError('الرجاء اختيار العقد أولاً / Please select a contract first.')
        if not openpyxl:
            raise UserError('مكتبة openpyxl غير مثبتة / openpyxl not installed.')

        try:
            file_content = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            raise UserError(f'فشل في قراءة الملف / Failed to read file: {str(e)}')

        if self.sheet_name and self.sheet_name.strip() in wb.sheetnames:
            ws = wb[self.sheet_name.strip()]
        else:
            ws = wb[wb.sheetnames[0]]

        max_col = ws.max_column or 0
        headers_combined, _ = self._build_merged_headers(ws, max_col)
        wb.close()

        col_map_model = self.env['delivery.column.map']
        existing_cols = {
            cm.excel_column.strip().lower()
            for cm in col_map_model.search([('contract_id', '=', self.contract_id.id)])
        }

        default_map = DEFAULT_COLUMN_MAPPING
        field_reverse = {}
        for field_name, aliases in default_map.items():
            alias_list = aliases if isinstance(aliases, list) else [aliases]
            for alias in alias_list:
                field_reverse[str(alias).strip().lower()] = field_name

        new_maps = []
        for seq, header_val in enumerate(headers_combined, 1):
            if not header_val:
                continue
            header_str = str(header_val).strip()
            if header_str.lower() in existing_cols:
                continue
            auto_field = field_reverse.get(header_str.lower())
            new_maps.append({
                'contract_id': self.contract_id.id,
                'sequence': seq * 10,
                'excel_column': header_str,
                'system_field': auto_field if auto_field and auto_field != 'row_num' else '_extra_',
                'custom_key': header_str.replace(' ', '_').lower() if not auto_field else False,
                'is_auto_detected': True,
            })

        if new_maps:
            col_map_model.create(new_maps)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'delivery.contract',
            'res_id': self.contract_id.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _build_merged_headers(self, ws, max_col):
        row1 = []
        row2 = []
        for col_idx in range(max_col):
            v1 = ws.cell(row=1, column=col_idx + 1).value
            v2 = ws.cell(row=2, column=col_idx + 1).value
            row1.append(str(v1).strip() if v1 is not None else None)
            row2.append(str(v2).strip() if v2 is not None else None)

        has_sub_headers = any(
            row2[i] is not None and row1[i] is None
            for i in range(max_col)
        )

        if has_sub_headers:
            merged = []
            for i in range(max_col):
                if row2[i] is not None:
                    merged.append(row2[i])
                elif row1[i] is not None:
                    merged.append(row1[i])
                else:
                    merged.append(None)
            data_start_row = 3
        else:
            merged = row1
            data_start_row = 2

        return merged, data_start_row

    def _match_rider(self, rider_model, rider_account_id, rider_phone, plate_number,
                     rider_user, rider_name, company_id):
        """مطابقة المندوب بالترتيب التالي:
        1. ID مباشر في نفس الشركة
        2. ID مباشر في أي شركة
        3. فريلانسر — بيرنته له نفس الـ ID (sub_rider_ids)
        4. رقم الجوال
        5. رقم اللوحة
        6. الاسم
        """
        matched = False

        # 1 & 2: ID مباشر
        if rider_account_id:
            rider = rider_model.search([
                ('platform_account_id', '=', rider_account_id),
                ('primary_company_id', '=', company_id),
            ], limit=1)
            if not rider:
                rider = rider_model.search([
                    ('platform_account_id', '=', rider_account_id),
                ], limit=1)
            if rider:
                matched = rider.id

        # 3: فريلانسر — بيرنته له هذا الـ ID
        if not matched and rider_account_id:
            parent = rider_model.search([
                ('platform_account_id', '=', rider_account_id),
            ], limit=1)
            if parent and parent.sub_rider_ids:
                active_subs = parent.sub_rider_ids.filtered(
                    lambda r: r.active and (not company_id or r.primary_company_id.id == company_id)
                )
                if not active_subs:
                    active_subs = parent.sub_rider_ids.filtered(lambda r: r.active)
                if active_subs:
                    matched = active_subs[0].id

        # 4: رقم الجوال
        if not matched and rider_phone:
            rider = rider_model.search([
                ('phone', '=', rider_phone),
                ('primary_company_id', '=', company_id),
            ], limit=1)
            if not rider:
                rider = rider_model.search([('phone', '=', rider_phone)], limit=1)
            if rider:
                matched = rider.id

        # 5: رقم اللوحة
        if not matched and plate_number:
            plate_norm = plate_number.strip().replace(' ', '').upper()
            candidates = rider_model.search([
                ('license_plate', '!=', False),
                ('primary_company_id', '=', company_id),
            ], limit=0)
            for r in candidates:
                if r.license_plate and r.license_plate.strip().replace(' ', '').upper() == plate_norm:
                    matched = r.id
                    break
            if not matched:
                for r in rider_model.search([('license_plate', '!=', False)], limit=0):
                    if r.license_plate and r.license_plate.strip().replace(' ', '').upper() == plate_norm:
                        matched = r.id
                        break

        # 6: الاسم
        if not matched:
            for sname in [rider_user, rider_name]:
                if not sname or sname == '--':
                    continue
                rider = rider_model.search([
                    '|', ('name', '=', sname), ('name_ar', '=', sname),
                ], limit=1)
                if rider:
                    matched = rider.id
                    break

        return matched

    def _parse_sheet_rows(self, ws, col_map, extra_col_map, sheet_day_override,
                          rider_model, city_model, row_counter_start, company_id):
        """معالجة صفوف ورقة واحدة وإعادة (rows_to_create, total, valid, errors, error_lines, amount).
        sheet_day_override: رقم اليوم من اسم الورقة (للملفات متعددة الأوراق)، أو None.
        """
        max_col = ws.max_column or 0
        max_row = ws.max_row or 0
        headers_combined, data_start_row = self._build_merged_headers(ws, max_col)

        rows_to_create = []
        total = 0
        valid = 0
        errors = 0
        error_lines = []
        total_amount = 0.0

        SUMMARY_KW = ['اجمالي', 'اجمالى', 'المتبقي', 'المتبق', 'الإجمالي',
                      'Total', 'Remaining', 'Balance']

        for row_idx in range(data_start_row, max_row + 1):
            row_data = []
            for c in range(1, max_col + 1):
                row_data.append(ws.cell(row=row_idx, column=c).value)

            if all(v is None or str(v).strip() == '' for v in row_data):
                continue

            if self.import_type == 'fuel':
                first_text = ' '.join(str(v).strip() for v in row_data if v is not None and str(v).strip())
                if any(kw in first_text for kw in SUMMARY_KW):
                    continue

            total += 1
            row_num = row_counter_start + total

            def get_val(field):
                idx = col_map.get(field)
                if idx is not None and idx < len(row_data):
                    return row_data[idx]
                return None

            def get_str(field):
                v = get_val(field)
                s = str(v).strip() if v is not None else ''
                return s if s not in ('', 'None', 'nan') else False

            def get_float(field):
                v = get_val(field)
                if v is None:
                    return 0.0
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return 0.0

            row_vals = {
                'session_id': self.id,
                'row_number': row_num,
                'status': 'pending',
            }

            row_vals['order_id'] = get_str('order_id') or get_str('rider_account_id') or False
            row_vals['rider_phone'] = get_str('rider_phone') or False
            row_vals['rider_name'] = get_str('rider_name') or False
            row_vals['rider_user'] = get_str('rider_user') or False
            row_vals['city_name'] = get_str('city_name') or False
            row_vals['vehicle_type_company'] = get_str('vehicle_type_company') or False
            row_vals['vehicle_type_contract'] = get_str('vehicle_type_contract') or False
            row_vals['plate_number'] = get_str('plate_number') or False

            row_vals['platform_target'] = get_float('platform_target')
            row_vals['company_target_val'] = get_float('company_target_val')
            row_vals['accepted_tasks'] = int(get_float('accepted_tasks'))
            row_vals['delivered_tasks'] = int(get_float('delivered_tasks'))
            row_vals['large_orders_completed'] = int(get_float('large_orders_completed'))
            row_vals['cancelled_tasks'] = int(get_float('cancelled_tasks'))
            row_vals['rejected_tasks'] = int(get_float('rejected_tasks'))
            row_vals['driver_rejected'] = int(get_float('driver_rejected'))
            row_vals['auto_rejected'] = int(get_float('auto_rejected'))
            # ساعات الاتصال — يدعم HH:MM وصيغة عربية وعدد عشري
            row_vals['online_hours'] = self._parse_online_hours(get_val('online_hours'))
            row_vals['peak_hours'] = self._parse_online_hours(get_val('peak_hours'))
            row_vals['ontime_rate'] = get_float('ontime_rate')
            row_vals['large_ontime_rate'] = get_float('large_ontime_rate')
            row_vals['avg_delivery_duration'] = get_float('avg_delivery_duration')
            row_vals['over_55min_rate'] = get_float('over_55min_rate')
            row_vals['late_tasks'] = int(get_float('late_tasks'))
            row_vals['very_late_tasks'] = int(get_float('very_late_tasks'))
            row_vals['fuel'] = get_float('fuel')
            row_vals['on_time_deliveries'] = int(get_float('on_time_deliveries'))
            row_vals['advance_deliveries'] = int(get_float('advance_deliveries'))
            row_vals['valid_days'] = int(get_float('valid_days'))
            row_vals['capacity_incentive'] = get_float('capacity_incentive')
            row_vals['experience_incentive'] = get_float('experience_incentive')
            row_vals['subsidy'] = get_float('subsidy')
            row_vals['dxg'] = get_float('dxg')
            row_vals['tips_excl_vat'] = get_float('tips_excl_vat')
            row_vals['other_activities'] = get_float('other_activities')
            row_vals['deductions'] = get_float('deductions')
            row_vals['food_damage_compensation'] = get_float('food_damage_compensation')
            row_vals['other_adjustment'] = get_float('other_adjustment')

            extra = {}
            for ecol_idx, ekey in extra_col_map.items():
                if ecol_idx < len(row_data):
                    ev = row_data[ecol_idx]
                    if ev is not None and str(ev).strip() not in ('', 'None'):
                        extra[ekey] = str(ev).strip()
            if extra:
                row_vals['extra_data'] = json.dumps(extra, ensure_ascii=False)

            # التاريخ — يدعم YYYYMMDD (كيتا) وdatetime ونص ورقم اليوم من الورقة
            date_from_col = self._parse_date_val(get_val('order_date'))
            if date_from_col:
                row_vals['order_date'] = date_from_col
            elif sheet_day_override is not None:
                row_vals['order_date'] = self._parse_date_val(None, sheet_day_override=sheet_day_override) or False
            else:
                row_vals['order_date'] = False

            row_vals['distance'] = get_float('distance')
            amount_val = get_float('platform_amount')
            row_vals['platform_amount'] = amount_val
            total_amount += amount_val

            # ─── مطابقة المندوب ───────────────────────────────────────────────
            rider_account_id = get_str('rider_account_id') or False
            matched_rider = self._match_rider(
                rider_model,
                rider_account_id,
                row_vals.get('rider_phone'),
                row_vals.get('plate_number'),
                row_vals.get('rider_user'),
                row_vals.get('rider_name'),
                company_id,
            )

            # إنشاء رايدر تلقائياً إذا لم يُعثر عليه والخيار مفعّل
            if not matched_rider and self.auto_create_riders and rider_account_id:
                try:
                    rider_name_val = (
                        row_vals.get('rider_name') or
                        row_vals.get('rider_user') or
                        rider_account_id
                    )
                    new_rider = rider_model.create({
                        'name': rider_name_val,
                        'platform_account_id': rider_account_id,
                        'primary_company_id': company_id or False,
                        'phone': row_vals.get('rider_phone') or False,
                        'license_plate': row_vals.get('plate_number') or False,
                        'active': True,
                    })
                    matched_rider = new_rider.id
                except Exception as e:
                    _logger.warning(
                        '3PL Import: auto-create rider failed for account %s: %s',
                        rider_account_id, str(e)
                    )

            row_vals['rider_id'] = matched_rider

            # ─── مطابقة المدينة ───────────────────────────────────────────────
            matched_city = False
            if row_vals.get('city_name'):
                city = city_model.search([
                    '|',
                    ('name', 'ilike', row_vals['city_name']),
                    ('name_ar', 'ilike', row_vals['city_name']),
                ], limit=1)
                if city:
                    matched_city = city.id
            row_vals['city_id'] = matched_city

            # ─── الحالة ───────────────────────────────────────────────────────
            row_errors = []
            if not matched_rider:
                identifier = (rider_account_id or row_vals.get('rider_phone')
                              or row_vals.get('rider_user') or row_vals.get('rider_name')
                              or f'Row {row_num}')
                hint = ' [فعّل "إنشاء رايدرز تلقائياً" لإضافتهم]' if not self.auto_create_riders else ''
                row_errors.append(f'مندوب غير معروف / Rider not found: {identifier}{hint}')

            if row_errors:
                row_vals['status'] = 'error'
                row_vals['error_message'] = '; '.join(row_errors)
                errors += 1
                error_lines.append(f"صف {row_num}: {'; '.join(row_errors)}")
            else:
                row_vals['status'] = 'valid'
                valid += 1

            rows_to_create.append(row_vals)

        return rows_to_create, total, valid, errors, error_lines, total_amount

    def write(self, vals):
        result = super().write(vals)
        # معالجة الملف تلقائياً عند رفعه إذا كانت جميع الحقول المطلوبة مكتملة
        if 'file_data' in vals and vals.get('file_data'):
            for rec in self:
                if (rec.status == 'pending'
                        and rec.file_data
                        and rec.period_start
                        and rec.period_end
                        and rec.company_id):
                    try:
                        rec.action_parse_file()
                    except Exception as e:
                        _logger.warning(
                            '3PL Import: auto-parse on upload failed for session %s: %s',
                            rec.id, str(e)
                        )
        return result

    def action_parse_file(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError('الرجاء رفع ملف Excel أولاً. / Please upload an Excel file first.')
        if not openpyxl:
            raise UserError('مكتبة openpyxl غير مثبتة. الرجاء تثبيتها: pip install openpyxl')

        self.row_ids.unlink()

        try:
            file_content = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            self.write({
                'status': 'failed',
                'error_log': f'فشل في قراءة الملف / Failed to read file: {str(e)}',
            })
            return

        # ─── تحديد الأوراق المراد قراءتها ───────────────────────────────────
        if self.import_all_sheets:
            sheets_to_read = wb.sheetnames  # كل الأوراق
        elif self.sheet_name and self.sheet_name.strip():
            target = self.sheet_name.strip()
            if target not in wb.sheetnames:
                available = ', '.join(wb.sheetnames)
                wb.close()
                self.write({
                    'status': 'failed',
                    'error_log': (
                        f'الورقة "{target}" غير موجودة.\n'
                        f'الأوراق المتاحة: {available}\n\n'
                        f'Sheet "{target}" not found. Available: {available}'
                    ),
                })
                return
            sheets_to_read = [target]
        else:
            sheets_to_read = [wb.sheetnames[0]]

        rider_model = self.env['delivery.rider']
        city_model = self.env['delivery.city']
        import_row_model = self.env['delivery.import.row']
        company_id = self.company_id.id if self.company_id else False

        # ─── كشف الأعمدة من الورقة الأولى في القائمة ────────────────────────
        first_ws = wb[sheets_to_read[0]]
        first_max_col = first_ws.max_column or 0
        first_max_row = first_ws.max_row or 0

        if first_max_row < 2 or first_max_col < 1:
            wb.close()
            self.write({
                'status': 'failed',
                'error_log': (
                    f'الورقة "{sheets_to_read[0]}" فارغة أو لا تحتوي على بيانات كافية.\n'
                    f'Sheet "{sheets_to_read[0]}" is empty or has insufficient data.'
                ),
            })
            return

        headers_combined, _ = self._build_merged_headers(first_ws, first_max_col)
        col_map, extra_col_map, unmapped_headers = self._detect_columns(headers_combined)

        if not col_map and not extra_col_map:
            header_names = [h for h in headers_combined if h]
            wb.close()
            self.write({
                'status': 'failed',
                'error_log': (
                    f'الورقة المستخدمة: {sheets_to_read[0]}\n'
                    f'لم يتم التعرف على أي عمود. / No columns were detected.\n'
                    f'أعمدة الملف / File headers: {", ".join(header_names)}\n\n'
                    f'استخدم زر "كشف وإعداد الأعمدة" أولاً لربط أعمدة الملف بحقول النظام.\n'
                    f'Use "Detect & Setup Columns" button first to map file columns to system fields.'
                ),
            })
            return

        # ─── قراءة جميع الأوراق ──────────────────────────────────────────────
        all_rows = []
        total = valid = errors = 0
        all_error_lines = []
        total_amount = 0.0
        sheets_info = []

        for sheet_name in sheets_to_read:
            ws = wb[sheet_name]

            # تحديد يوم الورقة — إذا كان اسم الورقة رقماً (1-31) نستخدمه كرقم اليوم
            sheet_day = None
            if self.import_all_sheets:
                try:
                    day_num = int(sheet_name.strip())
                    if 1 <= day_num <= 31:
                        sheet_day = day_num
                except (ValueError, TypeError):
                    pass

            rows, t, v, e, el, amt = self._parse_sheet_rows(
                ws, col_map, extra_col_map,
                sheet_day_override=sheet_day,
                rider_model=rider_model,
                city_model=city_model,
                row_counter_start=total,
                company_id=company_id,
            )
            all_rows.extend(rows)
            total += t
            valid += v
            errors += e
            all_error_lines.extend(el)
            total_amount += amt
            if t:
                sheets_info.append(f'  📄 {sheet_name}: {t} صف / rows ({v} صحيح / valid)')

        wb.close()

        if all_rows:
            import_row_model.create(all_rows)

        # ─── بناء سجل التشغيل ────────────────────────────────────────────────
        sheet_used = ', '.join(sheets_to_read) if len(sheets_to_read) <= 5 else f'{sheets_to_read[0]} … {sheets_to_read[-1]} ({len(sheets_to_read)} أوراق)'
        detected_cols = [f"  ✅ {k} → Col {v+1}: {headers_combined[v]}"
                         for k, v in sorted(col_map.items(), key=lambda x: x[1])]
        extra_detected = [f"  📦 extra[{ekey}] → Col {eidx+1}: {headers_combined[eidx]}"
                          for eidx, ekey in sorted(extra_col_map.items())]

        log_parts = [
            f'الأوراق المستخدمة / Sheets: {sheet_used}',
            f'نوع الاستيراد / Import Type: {self.import_type}',
            f'إجمالي أعمدة الملف / Total columns: {first_max_col}',
            '',
        ]
        if sheets_info:
            log_parts += ['تفاصيل الأوراق:'] + sheets_info + ['']

        log_parts += [f'الأعمدة المكتشفة والمربوطة / Detected system columns ({len(col_map)}):'] + detected_cols

        if extra_detected:
            log_parts += ['', f'أعمدة إضافية / Extra columns ({len(extra_col_map)}):'] + extra_detected

        if unmapped_headers:
            log_parts += [
                '',
                f'⚠️ أعمدة لم يتم ربطها / Unassigned columns ({len(unmapped_headers)}):',
                '   (محفوظة في extra_data — استخدم "كشف وإعداد الأعمدة" لربطها)',
            ]
            for uh in unmapped_headers:
                log_parts.append(f'   - {uh}')

        if all_error_lines:
            log_parts += ['', f'أخطاء / Errors ({errors}):'] + all_error_lines[:100]
            if len(all_error_lines) > 100:
                log_parts.append(f'... و {len(all_error_lines) - 100} أخطاء أخرى')
        else:
            log_parts += ['', 'لا توجد أخطاء / No errors found.']

        self.write({
            'total_rows': total,
            'valid_rows': valid,
            'error_rows': errors,
            'total_amount': total_amount,
            'error_log': '\n'.join(log_parts),
        })

        self._auto_create_performance()

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'delivery.import.session',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def _auto_create_performance(self):
        perf_model = self.env['delivery.daily.performance']
        monthly_model = self.env['delivery.monthly.performance']
        target_model = self.env['delivery.company.target']
        daily_created = 0
        daily_updated = 0
        perf_errors = []
        import_type = self.import_type or 'daily'

        valid_rows = self.row_ids.filtered(lambda r: r.rider_id)
        _logger.info('3PL Import: _auto_create_performance session %s, type=%s, %d matched rows',
                     self.id, import_type, len(valid_rows))

        if not valid_rows:
            _logger.warning('3PL Import: No rows with matched riders for session %s', self.id)
            return

        # ─── حل الفرع على مستوى الجلسة ───────────────────────────────────────────
        # ترتيب الأولوية:
        #   1. فرع الجلسة المحدد مباشرة
        #   2. فرع العقد المرتبط بالجلسة
        #   3. الفرع الوحيد للشركة (إذا كانت الشركة تملك فرعاً واحداً فقط)
        #   4. فرع كل مندوب (يُحسب لاحقاً صف بصف)
        session_branch_id = self.branch_id.id if self.branch_id else False
        if not session_branch_id and self.contract_id and self.contract_id.branch_id:
            session_branch_id = self.contract_id.branch_id.id
            _logger.info('3PL Import: session %s — using contract branch %s as fallback',
                         self.id, session_branch_id)
        if not session_branch_id and self.company_id:
            company_branches = self.env['delivery.company.branch'].search([
                ('company_id', '=', self.company_id.id)
            ])
            if len(company_branches) == 1:
                session_branch_id = company_branches.id
                _logger.info('3PL Import: session %s — using sole company branch %s as fallback',
                             self.id, session_branch_id)
            elif len(company_branches) == 0:
                _logger.warning('3PL Import: session %s — company has NO branches!', self.id)
            else:
                _logger.warning(
                    '3PL Import: session %s has no branch_id — will fallback to each rider\'s branch',
                    self.id)

        rider_daily_data = {}

        period_ref = self.period_start
        period_month = period_ref.month if period_ref else fields.Date.today().month
        period_year = period_ref.year if period_ref else fields.Date.today().year

        # ─── Monthly Import ───────────────────────────────────────────────────────
        if import_type == 'monthly':
            monthly_created = 0
            monthly_updated = 0
            for row in valid_rows:
                try:
                    with self.env.cr.savepoint():
                        effective_branch_id = session_branch_id or (
                            row.rider_id.branch_id.id if row.rider_id.branch_id else False)
                        if not effective_branch_id:
                            perf_errors.append(
                                f'Row {row.row_number}: لا يوجد فرع للجلسة أو للمندوب / '
                                f'No branch on session or rider ({row.rider_id.name})')
                            continue
                        monthly_vals = {
                            'branch_id': effective_branch_id,
                            'rider_id': row.rider_id.id,
                            'period_month': period_month,
                            'period_year': period_year,
                            'delivered_orders': int(row.delivered_tasks or 0),
                            'accepted_orders': int(row.accepted_tasks or 0),
                            'cancelled_orders': int(row.cancelled_tasks or 0),
                            'valid_hours': float(row.online_hours or 0.0),
                            'valid_days': int(row.valid_days or 0),
                            'capacity_incentive': float(row.capacity_incentive or 0.0),
                            'experience_incentive': float(row.experience_incentive or 0.0),
                            'subsidy': float(row.subsidy or 0.0),
                            'dxg': float(row.dxg or 0.0),
                            'tips_excl_vat': float(row.tips_excl_vat or 0.0),
                            'other_activities': float(row.other_activities or 0.0),
                            'deductions': float(row.deductions or 0.0),
                            'food_damage_compensation': float(row.food_damage_compensation or 0.0),
                            'other_adjustment': float(row.other_adjustment or 0.0),
                            'on_time_deliveries': int(row.on_time_deliveries or 0),
                            'advance_deliveries': int(row.advance_deliveries or 0),
                        }
                        # احسب الصلاحية مباشرة من valid_days المستوردة
                        import calendar as _cal
                        _days_in_month = _cal.monthrange(period_year, period_month)[1]
                        _tmp_rec = monthly_model.new(monthly_vals)
                        _criteria = _tmp_rec._get_active_validity_criteria()
                        _is_valid, _reason = _tmp_rec._check_monthly_validity(
                            int(row.valid_days or 0), _days_in_month, _criteria)
                        monthly_vals['is_valid'] = _is_valid
                        monthly_vals['validity_reason'] = _reason or ''

                        existing_m = monthly_model.search([
                            ('rider_id', '=', row.rider_id.id),
                            ('period_month', '=', period_month),
                            ('period_year', '=', period_year),
                            ('branch_id', '=', effective_branch_id),
                        ], limit=1)
                        if existing_m:
                            existing_m.write(monthly_vals)
                            monthly_updated += 1
                        else:
                            monthly_model.create(monthly_vals)
                            monthly_created += 1
                        rider_daily_data[(row.rider_id.id, period_month, period_year)] = {
                            'rider_id': row.rider_id.id,
                            'branch_id': effective_branch_id,
                            'month': period_month,
                            'year': period_year,
                            'delivered': int(row.delivered_tasks or 0),
                            'accepted': int(row.accepted_tasks or 0),
                            'cancelled': int(row.cancelled_tasks or 0),
                            'online_hours': float(row.online_hours or 0.0),
                            'fuel': float(row.fuel or 0.0),
                            'days': int(row.valid_days or 0),
                        }
                except Exception as e:
                    _logger.error('3PL Import: Error monthly row %s: %s', row.row_number, str(e))
                    perf_errors.append(f'Row {row.row_number}: {str(e)}')

            summary_parts = [
                f'\n\n--- نتائج الإنشاء التلقائي (استيراد شهري) ---',
                f'أداء شهري / Monthly: إنشاء {monthly_created} / تحديث {monthly_updated}',
            ]
            if perf_errors:
                summary_parts.append(f'أخطاء ({len(perf_errors)}):')
                summary_parts.extend(perf_errors[:20])
            current_log = self.error_log or ''
            self.write({'error_log': current_log + '\n'.join(summary_parts)})
            return

        # ─── Fuel Import ─────────────────────────────────────────────────────────
        if import_type == 'fuel':
            fuel_created = 0
            fuel_updated = 0
            fuel_date = self.period_start or fields.Date.today()
            for row in valid_rows:
                try:
                    with self.env.cr.savepoint():
                        effective_branch_id = session_branch_id or (
                            row.rider_id.branch_id.id if row.rider_id.branch_id else False)
                        if not effective_branch_id:
                            perf_errors.append(
                                f'Row {row.row_number}: لا يوجد فرع للجلسة أو للمندوب / '
                                f'No branch on session or rider ({row.rider_id.name})')
                            continue
                        fuel_amount = float(row.fuel or 0.0) or float(row.platform_amount or 0.0)
                        if not fuel_amount:
                            continue
                        existing_perf = perf_model.search([
                            ('rider_id', '=', row.rider_id.id),
                            ('date', '=', fuel_date),
                            ('branch_id', '=', effective_branch_id),
                        ], limit=1)
                        if existing_perf:
                            existing_perf.write({'fuel': fuel_amount})
                            fuel_updated += 1
                        else:
                            perf_model.create({
                                'branch_id': effective_branch_id,
                                'rider_id': row.rider_id.id,
                                'date': fuel_date,
                                'import_session_id': self.id,
                                'fuel': fuel_amount,
                                'license_plate': row.plate_number or False,
                            })
                            fuel_created += 1
                        if row.plate_number and not row.rider_id.license_plate:
                            row.rider_id.write({'license_plate': row.plate_number})
                        rider_daily_data[(row.rider_id.id, fuel_date.month, fuel_date.year)] = {
                            'rider_id': row.rider_id.id,
                            'branch_id': effective_branch_id,
                            'month': fuel_date.month,
                            'year': fuel_date.year,
                            'delivered': 0, 'accepted': 0, 'cancelled': 0,
                            'online_hours': 0.0, 'fuel': fuel_amount, 'days': 0,
                        }
                except Exception as e:
                    _logger.error('3PL Import: Error fuel row %s: %s', row.row_number, str(e))
                    perf_errors.append(f'Row {row.row_number}: {str(e)}')

            summary_parts = [
                f'\n\n--- نتائج الاستيراد (تقرير بنزين يومي) ---',
                f'تاريخ البنزين / Fuel Date: {fuel_date}',
                f'تحديث / Updated: {fuel_updated}  |  إنشاء جديد / Created: {fuel_created}',
            ]
            if perf_errors:
                summary_parts.append(f'أخطاء ({len(perf_errors)}):')
                summary_parts.extend(perf_errors[:20])
            current_log = self.error_log or ''
            self.write({'error_log': current_log + '\n'.join(summary_parts)})
            return

        # ─── Daily Import ─────────────────────────────────────────────────────────
        for row in valid_rows:
            try:
                with self.env.cr.savepoint():
                    effective_branch_id = session_branch_id or (
                        row.rider_id.branch_id.id if row.rider_id.branch_id else False)
                    if not effective_branch_id:
                        perf_errors.append(
                            f'Row {row.row_number}: لا يوجد فرع للجلسة أو للمندوب / '
                            f'No branch on session or rider ({row.rider_id.name})')
                        continue

                    perf_date = row.order_date or self.period_start
                    if not perf_date:
                        perf_date = fields.Date.context_today(self)

                    existing = perf_model.search([
                        ('rider_id', '=', row.rider_id.id),
                        ('date', '=', perf_date),
                        ('branch_id', '=', effective_branch_id),
                    ], limit=1)

                    perf_vals = {
                        'branch_id': effective_branch_id,
                        'rider_id': row.rider_id.id,
                        'date': perf_date,
                        'import_session_id': self.id,
                        'platform_account_id': row.order_id or False,
                        'account_name': row.rider_user or row.rider_name or False,
                        'vehicle_type_company': self._normalize_vehicle_type(row.vehicle_type_company),
                        'vehicle_type_contract': self._normalize_vehicle_type(row.vehicle_type_contract),
                        'license_plate': row.plate_number or False,
                        'platform_target': int(row.platform_target or 0),
                        'accepted_orders': int(row.accepted_tasks or 0),
                        'delivered_orders': int(row.delivered_tasks or 0),
                        'large_orders_completed': int(row.large_orders_completed or 0),
                        'cancelled_orders': int(row.cancelled_tasks or 0),
                        'rejected_orders': int(row.rejected_tasks or 0),
                        'total_online_hours': float(row.online_hours or 0.0),
                        'peak_hours': float(row.peak_hours or 0.0),
                        'on_time_deliveries': int(row.on_time_deliveries or 0),
                        'advance_deliveries': int(row.advance_deliveries or 0),
                    }

                    if existing:
                        existing.write(perf_vals)
                        daily_updated += 1
                    else:
                        perf_model.create(perf_vals)
                        daily_created += 1

                    p_month = perf_date.month if hasattr(perf_date, 'month') else int(str(perf_date).split('-')[1])
                    p_year = perf_date.year if hasattr(perf_date, 'year') else int(str(perf_date).split('-')[0])
                    rider_key = (row.rider_id.id, p_month, p_year)
                    if rider_key not in rider_daily_data:
                        rider_daily_data[rider_key] = {
                            'rider_id': row.rider_id.id,
                            'branch_id': effective_branch_id,
                            'month': p_month, 'year': p_year,
                            'delivered': 0, 'accepted': 0, 'cancelled': 0,
                            'online_hours': 0.0, 'days': 0, 'ontime': 0, 'fuel': 0.0,
                        }
                    rd = rider_daily_data[rider_key]
                    rd['delivered'] += int(row.delivered_tasks or 0)
                    rd['accepted'] += int(row.accepted_tasks or 0)
                    rd['cancelled'] += int(row.cancelled_tasks or 0)
                    rd['online_hours'] += float(row.online_hours or 0.0)
                    rd['fuel'] += float(row.fuel or 0.0)
                    rd['days'] += 1

                    rider_updates = {}
                    if row.order_id and not row.rider_id.platform_account_id:
                        rider_updates['platform_account_id'] = row.order_id
                    if row.vehicle_type_contract:
                        vt = self._normalize_vehicle_type(row.vehicle_type_contract)
                        if vt and not row.rider_id.vehicle_type:
                            rider_updates['vehicle_type'] = vt
                    if row.plate_number and not row.rider_id.license_plate:
                        rider_updates['license_plate'] = row.plate_number
                    if rider_updates:
                        row.rider_id.write(rider_updates)

            except Exception as e:
                _logger.error('3PL Import: Error daily perf row %s: %s', row.row_number, str(e))
                perf_errors.append(f'Row {row.row_number}: {str(e)}')

        # ─── Monthly Aggregation (from daily) ────────────────────────────────────
        monthly_created = 0
        monthly_updated = 0
        for key, rd in rider_daily_data.items():
            try:
                with self.env.cr.savepoint():
                    agg_branch_id = rd.get('branch_id') or session_branch_id
                    if not agg_branch_id:
                        perf_errors.append(
                            f'Monthly rider {rd["rider_id"]}: لا يوجد فرع / No branch — skip aggregation')
                        continue
                    existing_m = monthly_model.search([
                        ('rider_id', '=', rd['rider_id']),
                        ('period_month', '=', rd['month']),
                        ('period_year', '=', rd['year']),
                        ('branch_id', '=', agg_branch_id),
                    ], limit=1)
                    monthly_vals = {
                        'branch_id': agg_branch_id,
                        'rider_id': rd['rider_id'],
                        'period_month': rd['month'],
                        'period_year': rd['year'],
                        'delivered_orders': rd['delivered'],
                        'accepted_orders': rd['accepted'],
                        'cancelled_orders': rd['cancelled'],
                        'valid_hours': rd['online_hours'],
                        'valid_days': rd['days'],
                    }
                    if existing_m:
                        existing_m.write(monthly_vals)
                        monthly_rec = existing_m
                        monthly_updated += 1
                    else:
                        monthly_rec = monthly_model.create(monthly_vals)
                        monthly_created += 1
                    # احسب الصلاحية من السجلات اليومية تلقائياً
                    # (يُعيد حساب valid_days من is_valid_day اليومي ويضبط is_valid)
                    monthly_rec._aggregate_from_daily_records()
            except Exception as e:
                _logger.error('3PL Import: Error monthly perf rider %s: %s', rd['rider_id'], str(e))
                perf_errors.append(f'Monthly rider {rd["rider_id"]}: {str(e)}')

        # ─── Company Target ───────────────────────────────────────────────────────
        target_created = 0
        if rider_daily_data:
            sample = list(rider_daily_data.values())[0]
            t_month = sample['month']
            t_year = sample['year']
            t_branch_id = session_branch_id or sample.get('branch_id') or False
            company_id = self.company_id.id if self.company_id else False
            if company_id and t_branch_id:
                try:
                    with self.env.cr.savepoint():
                        existing_t = target_model.search([
                            ('company_id', '=', company_id),
                            ('branch_id', '=', t_branch_id),
                            ('month', '=', t_month),
                            ('year', '=', t_year),
                        ], limit=1)
                        if not existing_t:
                            target_model.create({
                                'company_id': company_id,
                                'branch_id': t_branch_id,
                                'month': t_month,
                                'year': t_year,
                            })
                            target_created += 1
                except Exception as e:
                    _logger.error('3PL Import: Error creating target: %s', str(e))
                    perf_errors.append(f'Target: {str(e)}')

        # ─── Settlement Auto-Create ───────────────────────────────────────────────
        settlement_created = False
        if rider_daily_data and self.contract_id and self.branch_id:
            try:
                with self.env.cr.savepoint():
                    sample = list(rider_daily_data.values())[0]
                    s_month = sample['month']
                    s_year = sample['year']
                    import datetime
                    period_start = datetime.date(s_year, s_month, 1)
                    period_end = (datetime.date(s_year, 12, 31) if s_month == 12
                                  else datetime.date(s_year, s_month + 1, 1) - datetime.timedelta(days=1))

                    settlement_model = self.env['delivery.settlement']
                    existing_settlement = settlement_model.search([
                        ('branch_id', '=', self.branch_id.id),
                        ('company_id', '=', self.company_id.id),
                        ('period_start', '=', period_start),
                    ], limit=1)

                    if not existing_settlement:
                        settlement = settlement_model.create({
                            'company_id': self.company_id.id,
                            'branch_id': self.branch_id.id,
                            'contract_id': self.contract_id.id,
                            'settlement_number': f'SET-{self.branch_id.name}-{s_year}-{s_month:02d}',
                            'period_start': period_start,
                            'period_end': period_end,
                            'cycle': self.contract_id.settlement_cycle or 'monthly',
                        })
                        settlement_created = True
                        try:
                            with self.env.cr.savepoint():
                                settlement.action_compute_settlement()
                        except Exception as e2:
                            _logger.warning('3PL Import: Settlement compute failed: %s', str(e2))
                    elif existing_settlement.status == 'draft':
                        try:
                            with self.env.cr.savepoint():
                                existing_settlement.action_compute_settlement()
                        except Exception as e2:
                            _logger.warning('3PL Import: Settlement recompute failed: %s', str(e2))
            except Exception as e:
                _logger.error('3PL Import: Error creating settlement: %s', str(e))
                perf_errors.append(f'Settlement: {str(e)}')

        # ─── Summary Log ─────────────────────────────────────────────────────────
        summary_parts = [
            f'\n\n--- نتائج الإنشاء التلقائي ---',
        ]
        if not session_branch_id:
            summary_parts.append(
                '⚠️ الجلسة بدون فرع — تم استخدام فرع كل مندوب تلقائياً / '
                'Session has no branch — each rider\'s own branch was used as fallback')
        summary_parts += [
            f'أداء يومي / Daily: إنشاء {daily_created} / تحديث {daily_updated}',
            f'أداء شهري / Monthly: إنشاء {monthly_created} / تحديث {monthly_updated}',
        ]
        if target_created:
            summary_parts.append(f'تارجيت الشركة / Company Target: إنشاء {target_created}')
        if settlement_created:
            summary_parts.append(f'تسوية / Settlement: تم إنشاؤها تلقائياً')
        if perf_errors:
            summary_parts.append(f'أخطاء ({len(perf_errors)}):')
            summary_parts.extend(perf_errors[:20])

        current_log = self.error_log or ''
        self.write({'error_log': current_log + '\n'.join(summary_parts)})

    def action_validate(self):
        self.ensure_one()
        if self.status != 'pending':
            raise ValidationError('Only pending imports can be validated.')
        if not self.row_ids:
            raise ValidationError(
                'لا توجد صفوف للتحقق. الرجاء معالجة الملف أولاً بالضغط على "معالجة الملف".\n'
                'No rows to validate. Please process the file first by clicking "Process File".'
            )
        error_count = len(self.row_ids.filtered(lambda r: r.status == 'error'))
        valid_count = len(self.row_ids.filtered(lambda r: r.status == 'valid'))
        pending_count = len(self.row_ids.filtered(lambda r: r.status == 'pending'))
        self.write({
            'status': 'validated',
            'valid_rows': valid_count,
            'error_rows': error_count,
        })

    def _normalize_vehicle_type(self, raw_val):
        if not raw_val:
            return False
        val = str(raw_val).strip().lower()
        if val in ('car', 'private car', 'سيارة', 'car (private car)'):
            return 'car'
        if val in ('bike', 'motorcycle', 'motor bike', 'motorbike', 'دراجة', 'موتور',
                   'motorcycle / bike', 'دراجة نارية'):
            return 'motorcycle'
        return False

    def _parse_online_hours(self, val):
        """تحويل ساعات الاتصال من صيغ مختلفة إلى عدد عشري (ساعات).
        يدعم:
        - عدد عشري مباشر: 10.5 → 10.5
        - صيغة HH:MM (تويو): '10:07' → 10.117
        - صيغة HH:MM:SS: '10:07:30' → 10.125
        - صيغة عربية (كيتا): '7 د 40 ث' → 0.128
        - صيغة عربية بالساعات: '1 ساعة 7 د 40 ث' → 1.128
        - صيغة '0 ث' → 0.0
        """
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        if not s or s in ('-', 'N/A', 'n/a', ''):
            return 0.0
        # HH:MM أو HH:MM:SS
        if ':' in s:
            parts = s.split(':')
            try:
                h = int(parts[0])
                m = int(parts[1]) if len(parts) > 1 else 0
                sec = int(parts[2]) if len(parts) > 2 else 0
                return h + m / 60.0 + sec / 3600.0
            except (ValueError, IndexError):
                pass
        # صيغة عربية: X ساعة Y د Z ث
        hours = 0.0
        import re as _re
        m = _re.search(r'(\d+)\s*ساع', s)
        if m:
            hours += float(m.group(1))
        m = _re.search(r'(\d+)\s*د(?:قيقة|ق)?(?:\s|$)', s)
        if m:
            hours += float(m.group(1)) / 60.0
        m = _re.search(r'(\d+)\s*ث(?:انية|ان)?', s)
        if m:
            hours += float(m.group(1)) / 3600.0
        if hours > 0.0:
            return hours
        # محاولة تحويل عشري مباشر
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    def _parse_date_val(self, val, sheet_day_override=None):
        """تحويل قيم التاريخ من صيغ مختلفة إلى date.
        يدعم:
        - datetime/date مباشر
        - YYYYMMDD كعدد (كيتا): 20260401.0 → date(2026,4,1)
        - نص '2026-04-01'
        - sheet_day_override: رقم اليوم من اسم الورقة (تويو/هنقر)
        """
        from datetime import date as _date, datetime as _datetime
        if sheet_day_override is not None:
            # الورقة رقمها = يوم الشهر، نستخدم period_start كمرجع
            try:
                ps = self.period_start
                if ps:
                    return ps.replace(day=int(sheet_day_override))
            except (ValueError, AttributeError):
                pass
        if val is None:
            return False
        if isinstance(val, _datetime):
            return val.date()
        if isinstance(val, _date):
            return val
        if isinstance(val, (int, float)):
            date_int = int(val)
            if 19000101 < date_int <= 21991231:
                try:
                    y = date_int // 10000
                    mo = (date_int % 10000) // 100
                    d = date_int % 100
                    return _date(y, mo, d)
                except ValueError:
                    pass
        try:
            from datetime import datetime
            return datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass
        try:
            from datetime import datetime
            return datetime.strptime(str(val).strip(), '%d/%m/%Y').date()
        except (ValueError, TypeError):
            pass
        return False

    def action_confirm(self):
        self.ensure_one()
        if self.status != 'validated':
            raise ValidationError('Only validated imports can be confirmed.')
        self._auto_create_performance()
        self.write({'status': 'imported'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'delivery.import.session',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_mark_failed(self):
        self.ensure_one()
        self.write({'status': 'failed'})

    def action_reset_to_pending(self):
        self.ensure_one()
        self.row_ids.unlink()
        self.write({
            'status': 'pending',
            'total_rows': 0,
            'valid_rows': 0,
            'error_rows': 0,
            'total_amount': 0.0,
            'error_log': False,
        })

    def action_cancel_and_re_upload(self):
        """حذف بيانات الأداء المنشأة من هذه الجلسة وإعادتها لـ Pending لإعادة الرفع."""
        self.ensure_one()
        daily_perf = self.env['delivery.daily.performance'].search([
            ('import_session_id', '=', self.id)
        ])
        daily_count = len(daily_perf)
        daily_perf.unlink()

        self.row_ids.unlink()
        self.write({
            'status': 'pending',
            'total_rows': 0,
            'valid_rows': 0,
            'error_rows': 0,
            'total_amount': 0.0,
            'error_log': f'تم حذف {daily_count} سجل أداء يومي وإعادة الجلسة للرفع. / '
                         f'Deleted {daily_count} daily performance records and reset for re-upload.',
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'delivery.import.session',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }


class DeliveryImportRow(models.Model):
    _name = 'delivery.import.row'
    _description = 'Import Row'
    _order = 'row_number'

    session_id = fields.Many2one('delivery.import.session', string='Import Session', required=True, ondelete='cascade')
    row_number = fields.Integer(string='Row #', required=True)
    order_id = fields.Char(string='Order / Account ID')
    rider_phone = fields.Char(string='Rider Phone')
    rider_name = fields.Char(string='Rider Name (اسم الحساب)')
    rider_user = fields.Char(string='Account User (مستخدم الحساب)')
    rider_id = fields.Many2one('delivery.rider', string='Matched Rider')
    city_name = fields.Char(string='City Name')
    city_id = fields.Many2one('delivery.city', string='Matched City')
    order_date = fields.Date(string='Order Date')

    vehicle_type_company = fields.Char(string='Vehicle Type (Company)')
    vehicle_type_contract = fields.Char(string='Vehicle Type (Contract)')
    plate_number = fields.Char(string='Plate Number (رقم اللوحة)')

    platform_target = fields.Float(string='Platform Target (تارجت المنصة)', digits=(8, 2))
    company_target_val = fields.Float(string='Company Target (تارجت الشركة)', digits=(8, 2))

    accepted_tasks = fields.Integer(string='Accepted Tasks (المهام المقبولة)')
    delivered_tasks = fields.Integer(string='Delivered Tasks (المهام المسلّمة)')
    large_orders_completed = fields.Integer(string='Large Orders Completed')
    cancelled_tasks = fields.Integer(string='Cancelled Tasks (الملغاة)')
    rejected_tasks = fields.Integer(string='Rejected Tasks (المرفوضة)')
    driver_rejected = fields.Integer(string='Driver Rejected (رفض السائق)')
    auto_rejected = fields.Integer(string='Auto Rejected (رفض تلقائي)')

    online_hours = fields.Float(string='Online Hours (ساعات الاتصال)', digits=(8, 2))
    peak_hours = fields.Float(string='Peak Hours (ساعات الذروة)', digits=(8, 2))
    ontime_rate = fields.Float(string='On-time Rate (نسبة التسليم بالوقت)', digits=(5, 4))
    large_ontime_rate = fields.Float(string='Large Order On-time Rate', digits=(5, 4))
    avg_delivery_duration = fields.Float(string='Avg Delivery Duration (min)', digits=(8, 2))
    over_55min_rate = fields.Float(string='Over 55min Rate', digits=(5, 4))
    late_tasks = fields.Integer(string='Late Tasks (متأخرة)')
    very_late_tasks = fields.Integer(string='Very Late Tasks (متأخرة جداً)')
    on_time_deliveries = fields.Integer(string='On-Time Deliveries Count')
    advance_deliveries = fields.Integer(string='Advance Deliveries Count')

    fuel = fields.Float(string='Fuel (بنزين)', digits=(8, 2))
    valid_days = fields.Integer(string='Valid Days (أيام صالحة)')

    capacity_incentive = fields.Float(string='Capacity Incentive (حافز السعة)', digits=(12, 2))
    experience_incentive = fields.Float(string='Experience Incentive (حافز الخبرة)', digits=(12, 2))
    subsidy = fields.Float(string='Subsidy (إعانة)', digits=(12, 2))
    dxg = fields.Float(string='DXG', digits=(12, 2))
    tips_excl_vat = fields.Float(string='Tips excl. VAT (بقشيش)', digits=(12, 2))
    other_activities = fields.Float(string='Other Activities (أنشطة أخرى)', digits=(12, 2))
    deductions = fields.Float(string='Deductions (خصومات)', digits=(12, 2))
    food_damage_compensation = fields.Float(string='Food Damage (تعويض تلف طعام)', digits=(12, 2))
    other_adjustment = fields.Float(string='Other Adjustment (تعديل آخر)', digits=(12, 2))

    distance = fields.Float(string='Distance (km)', digits=(8, 2))
    platform_amount = fields.Float(string='Platform Amount', digits=(12, 2))
    calculated_amount = fields.Float(string='Calculated Amount', digits=(12, 2))
    variance = fields.Float(string='Variance', digits=(12, 2))

    extra_data = fields.Text(
        string='Extra Data (JSON)',
        help='بيانات الأعمدة الإضافية التي لم يتم ربطها بحقل محدد في النظام.\n'
             'Extra column data that was not mapped to a specific system field.')

    status = fields.Selection([
        ('pending', 'Pending'),
        ('valid', 'Valid'),
        ('error', 'Error'),
    ], string='Status', default='pending')
    error_message = fields.Char(string='Error Message')
