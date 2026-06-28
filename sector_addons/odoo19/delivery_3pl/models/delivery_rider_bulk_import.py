import base64
import io
import logging

from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


RIDER_COLUMN_ALIASES = {
    'name': [
        'Full Name (EN)', 'Full Name', 'Name', 'name', 'الاسم',
        'الاسم الكامل', 'اسم المندوب', 'الاسم بالانجليزي',
    ],
    'name_ar': [
        'Full Name (AR)', 'Arabic Name', 'name_ar', 'الاسم بالعربي',
        'الاسم العربي', 'الاسم بالعربية',
    ],
    'phone': [
        'Phone', 'Phone Number', 'phone', 'هاتف', 'رقم الجوال',
        'جوال', 'Mobile', 'رقم الهاتف',
    ],
    'national_id': [
        'National ID', 'Iqama', 'national_id', 'هوية', 'الهوية',
        'رقم الهوية', 'الإقامة', 'رقم الإقامة',
    ],
    'platform_account_id': [
        'Platform Account ID', 'Account ID', 'platform_account_id',
        'معرف الحساب', 'رقم الحساب', 'ID', 'Rider ID',
    ],
    'parent_rider': [
        'Parent Rider', 'Parent Rider (for subcontract)', 'parent_rider',
        'البيرنت', 'السائق الأصلي', 'الحساب الأصلي', 'Parent',
        'Parent Account', 'parent_account_id',
    ],
    'rider_type': [
        'Rider Type', 'Type', 'rider_type', 'نوع المندوب', 'نوع السائق',
        'الفئة',
    ],
    'primary_company': [
        'Primary Company', 'Company', 'primary_company', 'الشركة',
        'الشركة الأساسية', 'المنصة',
    ],
    'branch': [
        'Branch', 'branch', 'الفرع', 'فرع',
    ],
    'vehicle_type': [
        'Vehicle Type', 'vehicle_type', 'نوع المركبة', 'نوع السيارة',
    ],
    'license_plate': [
        'License Plate', 'Plate', 'plate_number', 'license_plate',
        'اللوحة', 'رقم اللوحة', 'لوحة',
    ],
    'status': [
        'Status', 'status', 'الحالة',
    ],
    'join_date': [
        'Join Date', 'join_date', 'تاريخ الانضمام', 'تاريخ',
    ],
    'work_start_date': [
        'Work Start Date', 'work_start_date', 'تاريخ بداية العمل',
    ],
    'vehicle_model': [
        'Vehicle Model', 'vehicle_model', 'موديل المركبة', 'موديل السيارة',
    ],
    'vehicle_ownership': [
        'Vehicle Ownership', 'vehicle_ownership', 'ملكية المركبة',
    ],
    'notes': [
        'Notes', 'notes', 'ملاحظات',
    ],
}


class DeliveryRiderBulkImport(models.TransientModel):
    _name = 'delivery.rider.bulk.import'
    _description = 'Rider Bulk Import Wizard (استيراد جماعي للسائقين)'

    file_data = fields.Binary(string='Excel File (ملف Excel)', required=True, attachment=False)
    file_name = fields.Char(string='File Name')
    sheet_name = fields.Char(
        string='Sheet Name (اسم الورقة)',
        help='اتركه فارغاً لقراءة الورقة الأولى. / Leave empty to read the first sheet.',
    )
    update_existing = fields.Boolean(
        string='Update Existing Riders (تحديث السائقين الموجودين)',
        default=True,
        help='إذا كان مفعلاً، يتم تحديث بيانات السائقين الموجودين. '
             'إذا كان معطلاً، يتم تخطي السائقين الموجودين.',
    )
    match_by = fields.Selection([
        ('platform_account_id', 'Platform Account ID (معرف الحساب) — موصى به'),
        ('phone', 'Phone Number (رقم الجوال)'),
        ('name', 'Full Name (الاسم الكامل)'),
    ], string='Match Existing Riders By (مطابقة السائقين الموجودين بـ)',
        default='platform_account_id', required=True,
        help='الحقل المستخدم لمطابقة سجلات السائقين الموجودة في النظام.',
    )
    default_rider_type_id = fields.Many2one(
        'delivery.rider.type',
        string='Default Rider Type (نوع المندوب الافتراضي)',
        help='يُستخدم إذا لم يكن عمود "نوع المندوب" موجوداً في الملف.',
    )
    default_company_id = fields.Many2one(
        'delivery.company',
        string='Default Company (الشركة الافتراضية)',
    )
    result_summary = fields.Text(string='Import Result', readonly=True)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('done', 'Done'),
    ], default='draft')

    def _detect_columns(self, headers):
        """Map column indices to field names using RIDER_COLUMN_ALIASES."""
        col_map = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            header_clean = str(header).strip().lower()
            for field_name, aliases in RIDER_COLUMN_ALIASES.items():
                if field_name in col_map:
                    continue
                for alias in aliases:
                    if header_clean == alias.strip().lower():
                        col_map[field_name] = col_idx
                        break
        return col_map

    def _normalize_vehicle_type(self, val):
        if not val:
            return False
        v = str(val).strip().lower()
        if v in ('car', 'private car', 'سيارة', 'سياره'):
            return 'car'
        if v in ('bike', 'motorcycle', 'دراجة', 'دراجه', 'موتور', 'موتوسيكل'):
            return 'motorcycle'
        return False

    def _normalize_status(self, val):
        if not val:
            return 'active'
        v = str(val).strip().lower()
        if v in ('active', 'نشط', 'نشيط', '1', 'yes', 'نعم'):
            return 'active'
        if v in ('inactive', 'غير نشط', 'غير فعال', '0', 'no', 'لا'):
            return 'inactive'
        if v in ('suspended', 'موقوف', 'معلق'):
            return 'suspended'
        return 'active'

    def _parse_date(self, val):
        if not val:
            return False
        if hasattr(val, 'date'):
            return val.date()
        if hasattr(val, 'strftime'):
            return val
        try:
            from datetime import datetime
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                try:
                    return datetime.strptime(str(val).strip(), fmt).date()
                except ValueError:
                    continue
        except Exception:
            pass
        return False

    def action_import(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError('مكتبة openpyxl غير مثبتة. الرجاء تثبيتها: pip install openpyxl')
        if not self.file_data:
            raise UserError('الرجاء رفع ملف Excel أولاً.')

        try:
            file_content = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            raise UserError(f'فشل في قراءة الملف: {str(e)}')

        if self.sheet_name and self.sheet_name.strip() in wb.sheetnames:
            ws = wb[self.sheet_name.strip()]
        else:
            ws = wb[wb.sheetnames[0]]

        max_col = ws.max_column or 0
        max_row = ws.max_row or 0

        if max_row < 2 or max_col < 1:
            wb.close()
            raise UserError('الملف فارغ أو لا يحتوي على بيانات كافية.')

        headers = [
            str(ws.cell(row=1, column=c + 1).value).strip()
            if ws.cell(row=1, column=c + 1).value is not None else ''
            for c in range(max_col)
        ]
        col_map = self._detect_columns(headers)

        if not col_map:
            wb.close()
            raise UserError(
                f'لم يتم التعرف على أي عمود في الملف.\n'
                f'أعمدة الملف: {", ".join(h for h in headers if h)}\n\n'
                f'الأعمدة المدعومة تشمل: Full Name (EN), Phone, Platform Account ID, '
                f'Parent Rider, Rider Type, ...'
            )

        def get_val(row_data, field):
            idx = col_map.get(field)
            if idx is not None and idx < len(row_data):
                v = row_data[idx]
                return str(v).strip() if v is not None and str(v).strip() not in ('', 'None') else False
            return False

        all_rows = []
        for row_idx in range(2, max_row + 1):
            row_data = [ws.cell(row=row_idx, column=c + 1).value for c in range(max_col)]
            if all(v is None or str(v).strip() == '' for v in row_data):
                continue
            all_rows.append((row_idx, row_data))
        wb.close()

        rider_model = self.env['delivery.rider']
        rider_type_model = self.env['delivery.rider.type']
        company_model = self.env['delivery.company']
        branch_model = self.env['delivery.company.branch']

        rider_type_cache = {}
        company_cache = {}
        branch_cache = {}

        created = 0
        updated = 0
        skipped = 0
        errors = []

        # ── مرحلة أولى: إنشاء أو تحديث السائقين بدون parent_rider_id ──────────
        rider_map = {}

        for row_idx, row_data in all_rows:
            try:
                name_en = get_val(row_data, 'name') or False
                name_ar = get_val(row_data, 'name_ar') or False
                phone = get_val(row_data, 'phone') or False
                account_id = get_val(row_data, 'platform_account_id') or False
                national_id = get_val(row_data, 'national_id') or False
                status_raw = get_val(row_data, 'status') or False
                join_date_raw = self._parse_date(
                    row_data[col_map['join_date']] if 'join_date' in col_map else None
                )
                work_start_raw = self._parse_date(
                    row_data[col_map['work_start_date']] if 'work_start_date' in col_map else None
                )
                vehicle_type_raw = get_val(row_data, 'vehicle_type') or False
                license_plate = get_val(row_data, 'license_plate') or False
                vehicle_model = get_val(row_data, 'vehicle_model') or False
                notes = get_val(row_data, 'notes') or False

                identifier = account_id or phone or name_en or name_ar
                if not identifier:
                    skipped += 1
                    continue

                # ── بحث عن السائق الموجود ─────────────────────────────────────
                existing = False
                if self.match_by == 'platform_account_id' and account_id:
                    existing = rider_model.search([('platform_account_id', '=', account_id)], limit=1)
                elif self.match_by == 'phone' and phone:
                    existing = rider_model.search([('phone', '=', phone)], limit=1)
                elif self.match_by == 'name':
                    if name_en:
                        existing = rider_model.search([('name', '=', name_en)], limit=1)
                    if not existing and name_ar:
                        existing = rider_model.search([('name_ar', '=', name_ar)], limit=1)

                # ── حل نوع المندوب ────────────────────────────────────────────
                rider_type_id = False
                rtype_raw = get_val(row_data, 'rider_type') or False
                if rtype_raw:
                    rtype_key = rtype_raw.lower()
                    if rtype_key not in rider_type_cache:
                        rt = rider_type_model.search([
                            '|', ('name', 'ilike', rtype_raw), ('name_ar', 'ilike', rtype_raw)
                        ], limit=1)
                        rider_type_cache[rtype_key] = rt.id if rt else False
                    rider_type_id = rider_type_cache[rtype_key]
                if not rider_type_id and self.default_rider_type_id:
                    rider_type_id = self.default_rider_type_id.id

                # ── حل الشركة ─────────────────────────────────────────────────
                company_id = False
                comp_raw = get_val(row_data, 'primary_company') or False
                if comp_raw:
                    comp_key = comp_raw.lower()
                    if comp_key not in company_cache:
                        co = company_model.search([
                            '|', ('name', 'ilike', comp_raw), ('name_ar', 'ilike', comp_raw)
                        ], limit=1)
                        company_cache[comp_key] = co.id if co else False
                    company_id = company_cache[comp_key]
                if not company_id and self.default_company_id:
                    company_id = self.default_company_id.id

                # ── حل الفرع ──────────────────────────────────────────────────
                branch_id = False
                branch_raw = get_val(row_data, 'branch') or False
                if branch_raw:
                    branch_key = (branch_raw.lower(), company_id or 0)
                    if branch_key not in branch_cache:
                        domain = [
                            '|', ('name', 'ilike', branch_raw), ('name_ar', 'ilike', branch_raw)
                        ]
                        if company_id:
                            domain = [('company_id', '=', company_id)] + domain
                        br = branch_model.search(domain, limit=1)
                        branch_cache[branch_key] = br.id if br else False
                    branch_id = branch_cache[branch_key]

                vals = {}
                if name_en:
                    vals['name'] = name_en
                if name_ar:
                    vals['name_ar'] = name_ar
                if phone:
                    vals['phone'] = phone
                if account_id:
                    vals['platform_account_id'] = account_id
                if national_id:
                    vals['national_id'] = national_id
                if rider_type_id:
                    vals['rider_type_id'] = rider_type_id
                if company_id:
                    vals['primary_company_id'] = company_id
                if branch_id:
                    vals['branch_id'] = branch_id
                if vehicle_type_raw:
                    vt = self._normalize_vehicle_type(vehicle_type_raw)
                    if vt:
                        vals['vehicle_type'] = vt
                if license_plate:
                    vals['license_plate'] = license_plate
                if vehicle_model:
                    vals['vehicle_model'] = vehicle_model
                if status_raw:
                    vals['status'] = self._normalize_status(status_raw)
                if join_date_raw:
                    vals['join_date'] = join_date_raw
                if work_start_raw:
                    vals['work_start_date'] = work_start_raw
                if notes:
                    vals['notes'] = notes

                if existing:
                    if self.update_existing:
                        existing.write(vals)
                        rider_rec = existing
                        updated += 1
                    else:
                        rider_rec = existing
                        skipped += 1
                else:
                    if 'phone' not in vals:
                        vals['phone'] = phone or identifier or 'N/A'
                    rider_rec = rider_model.create(vals)
                    created += 1

                lookup_key = identifier
                rider_map[lookup_key] = rider_rec.id
                if account_id:
                    rider_map[account_id] = rider_rec.id
                if phone:
                    rider_map[phone] = rider_rec.id
                if name_en:
                    rider_map[name_en.lower()] = rider_rec.id
                if name_ar:
                    rider_map[name_ar.lower()] = rider_rec.id

            except Exception as e:
                errors.append(f'صف {row_idx}: {str(e)}')
                _logger.error('RiderBulkImport: row %s error: %s', row_idx, str(e))

        # ── مرحلة ثانية: ربط parent_rider_id ────────────────────────────────
        parent_linked = 0
        parent_errors = []

        if 'parent_rider' in col_map:
            for row_idx, row_data in all_rows:
                parent_raw = get_val(row_data, 'parent_rider') or False
                if not parent_raw:
                    continue

                child_account = get_val(row_data, 'platform_account_id') or False
                child_phone = get_val(row_data, 'phone') or False
                child_name = get_val(row_data, 'name') or False

                child_id = (
                    rider_map.get(child_account)
                    or rider_map.get(child_phone)
                    or rider_map.get((child_name or '').lower())
                )
                if not child_id:
                    parent_errors.append(f'صف {row_idx}: لم يتم العثور على السائق الفرىلانسر.')
                    continue

                parent_id = (
                    rider_map.get(parent_raw)
                    or rider_map.get(parent_raw.lower())
                )
                if not parent_id:
                    parent_rec = rider_model.search([
                        '|', '|',
                        ('platform_account_id', '=', parent_raw),
                        ('name', '=', parent_raw),
                        ('name_ar', '=', parent_raw),
                    ], limit=1)
                    if not parent_rec:
                        parent_rec = rider_model.search([
                            '|',
                            ('name', 'ilike', parent_raw),
                            ('name_ar', 'ilike', parent_raw),
                        ], limit=1)
                    if parent_rec:
                        parent_id = parent_rec.id
                        rider_map[parent_raw] = parent_id
                        rider_map[parent_raw.lower()] = parent_id

                if parent_id:
                    try:
                        rider_model.browse(child_id).write({'parent_rider_id': parent_id})
                        parent_linked += 1
                    except Exception as e:
                        parent_errors.append(f'صف {row_idx}: {str(e)}')
                else:
                    parent_errors.append(
                        f'صف {row_idx}: لم يُوجَد البيرنت "{parent_raw}" في النظام أو الملف.'
                    )

        # ── ملخص ─────────────────────────────────────────────────────────────
        summary_lines = [
            '✅ اكتمل الاستيراد / Import Complete',
            f'',
            f'تم إنشاؤهم / Created : {created}',
            f'تم تحديثهم / Updated : {updated}',
            f'تم تخطيهم / Skipped  : {skipped}',
            f'ربط البيرنت / Parent Linked: {parent_linked}',
        ]
        if parent_errors:
            summary_lines += ['', f'⚠️ تحذيرات البيرنت ({len(parent_errors)}):']
            summary_lines += parent_errors[:30]
        if errors:
            summary_lines += ['', f'❌ أخطاء ({len(errors)}):']
            summary_lines += errors[:30]

        self.write({
            'result_summary': '\n'.join(summary_lines),
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_open_riders(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Riders (المناديب)',
            'res_model': 'delivery.rider',
            'view_mode': 'list,form',
            'target': 'current',
        }
